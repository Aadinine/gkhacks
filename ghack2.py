import random
import pandas as pd
from openpyxl import Workbook         
from openpyxl.styles import Font     



# Basic settings
days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
periods = [f'Period {i}' for i in range(1, 9)]
rooms = ['Room A', 'Room B', 'Room C']
lunch_period = 5

# 6 subjects
courses = [
    {'code': 'MATH', 'teacher': 'Dr. Smith'},
    {'code': 'PHY', 'teacher': 'Dr. Johnson'},
    {'code': 'CHEM', 'teacher': 'Dr. Brown'},
    {'code': 'BIO', 'teacher': 'Dr. Wilson'},
    {'code': 'CS', 'teacher': 'Dr. Davis'},
    {'code': 'ENG', 'teacher': 'Dr. Miller'}
]

def create_empty_timetable():
    timetable = {}  # Start with empty dictionary
    
    for day in days:
        timetable[day] = {}  # Create empty slot for each day
        for period in periods:
            timetable[day][period] = {}  # Create empty slot for each period
            for room in rooms:
                if period == f'Period {lunch_period}':
                    timetable[day][period][room] = 'LUNCH BREAK'
                else:
                    timetable[day][period][room] = 'EMPTY'
    
    return timetable
def print_complete_timetable(timetable):
    """Print the complete timetable for all days"""
    print("\n" + "=" * 90)
    print("📅 COMPLETE WEEKLY TIMETABLE")
    print("=" * 90)
    
    for day in days:
        print(f"\n{day.upper()}")
        print("-" * 90)
        print(f"{'Period':<12} {'Room A':<25} {'Room B':<25} {'Room C':<25}")
        print("-" * 90)
        
        for period in periods:
            room_a = timetable[day][period]['Room A']
            room_b = timetable[day][period]['Room B']
            room_c = timetable[day][period]['Room C']
            print(f"{period:<12} {room_a:<25} {room_b:<25} {room_c:<25}")

def check_teacher_free(teacher, day, period, timetable):
    """Check if a teacher is already teaching in this period"""
    # Check all rooms in this day and period
    for room in rooms:
        # If the teacher's name appears in any room this period, they're busy
        if teacher in timetable[day][period][room]:
            return False
    # If we didn't find the teacher in any room, they're free
    return True

def count_subject_in_day(timetable, day, subject_code):
    """Count how many times a subject appears in a single day"""
    count = 0
    for period in periods:
        for room in rooms:
            if subject_code in timetable[day][period][room]:
                count += 1
    return count

def fill_timetable_with_courses(timetable):
    """Fill the empty slots with courses"""
    
    # Loop through each day
    for day in days:
        print(f"  Filling {day}...")
        
        # Loop through each period
        for period in periods:
            # Skip lunch period (we don't add courses here)
            if period == f'Period {lunch_period}':
                continue
                
            # Loop through each room
            for room in rooms:
                # Check if this slot is empty
                if timetable[day][period][room] == 'EMPTY':
                    
                    # Find all courses that meet our rules
                    available_courses = []
                    for course in courses:
                        # Rule 1: Teacher must be free in this period
                        if not check_teacher_free(course['teacher'], day, period, timetable):
                            continue
                            
                        # Rule 2: Subject must not appear more than thrice today
                        subject_count = count_subject_in_day(timetable, day, course['code'])
                        if subject_count >= 3:
                            continue
                            
                        # If both rules pass, add to available courses
                        available_courses.append(course)
                    
                    # If we found available courses, pick one randomly
                    if available_courses:
                        chosen_course = random.choice(available_courses)
                        timetable[day][period][room] = f"{chosen_course['code']} - {chosen_course['teacher']}"
                    else:
                        # If no courses are available, mark as FREE
                        timetable[day][period][room] = 'FREE'
    
    return timetable


def print_subject_counts(timetable):
    """Print how many times each subject appears each day"""
    print("\n📊 SUBJECT COUNTS PER DAY (Max: 2 per day)")
    print("=" * 60)
    
    for day in days:
        print(f"\n{day}:")
        for course in courses:
            count = count_subject_in_day(timetable, day, course['code'])
            print(f"  {course['code']}: {count} times")



def save_timetable_to_excel(timetable, filename="school_timetable.xlsx"):
    """Save the timetable to an Excel file"""
    print(f"\n💾 Saving timetable to {filename}...")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    for day in days:
        ws = wb.create_sheet(title=day)
        
        headers = ['Period', 'Room A', 'Room B', 'Room C']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row, period in enumerate(periods, 2):
            ws.cell(row=row, column=1, value=period)
            ws.cell(row=row, column=2, value=timetable[day][period]['Room A'])
            ws.cell(row=row, column=3, value=timetable[day][period]['Room B'])
            ws.cell(row=row, column=4, value=timetable[day][period]['Room C'])
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
    
    wb.save(filename)
    print(f"✅ Timetable saved successfully as {filename}!")





# Create empty timetable
print("🔄 Step 1: Creating empty timetable...")
timetable = create_empty_timetable()

# Fill with courses
print("📚 Step 2: Filling timetable with courses...")
timetable = fill_timetable_with_courses(timetable)

# Print subject counts
print_subject_counts(timetable)


save_timetable_to_excel(timetable)
# Print the complete timetable
print("\n🖨️ Step 3: Printing timetable...")
print_complete_timetable(timetable)

# Print summary
print(f"\n📊 Summary:")
print(f"• {len(days)} days, {len(periods)} periods per day")
print(f"• {len(rooms)} classrooms, {len(courses)} subjects")
print(f"• Lunch break: Period {lunch_period}")
print(f"• Rule: Max 2 same subjects per day")